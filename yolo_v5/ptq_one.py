import argparse
import glob
import os
import shutil
from pathlib import Path
import openpyxl

import numpy as np
import torch
from tqdm import tqdm
from yolov5_model import Yolo_v5

import gol
import sys
sys.path.append('/workspace/I/wangyuanzhe0/Cam_Model_TF')
from extract_ratio import extract_ratio
from util import *
from module import *

from mapping import modify_quant_state

from utils.datasets import create_dataloader
from utils.general import (
    coco80_to_coco91_class, check_dataset, check_file, check_img_size, compute_loss, non_max_suppression, scale_coords, 
    xyxy2xywh, clip_coords, plot_images, xywh2xyxy, box_iou, output_to_target, ap_per_class, set_logging)
from utils.torch_utils import select_device, time_synchronized


def test(data,
         weights=None,
         batch_size=128,
         imgsz=640,
         conf_thres=0.001,
         iou_thres=0.6,  # for NMS
         save_json=False,
         single_cls=False,
         augment=False,
         verbose=False,
         model=None,
         dataloader=None,
         save_dir='',
         merge=False,
         save_txt=False,
         quantize=False,
         quant_type='INT',
         num_bits=8,
         e_bits=3
         ):
    
    # Initialize/load model and set device
    training = False
    if not training:  # called by train.py

        set_logging()
        device = select_device(opt.device, batch_size=batch_size)
        merge, save_txt = opt.merge, opt.save_txt  # use Merge NMS, save *.txt labels
        # Construct output directory
        if save_txt:
            out = Path('inference/output')
            if os.path.exists(out):
                shutil.rmtree(out)  # delete output folder
            os.makedirs(out)  # make new output folder

        # Remove previous
        for f in glob.glob(str(Path(save_dir) / 'test_batch*.jpg')):
            os.remove(f)

        # Load model
        #model = attempt_load(weights, map_location=device)  # load FP32 model
        imgsz = check_img_size(imgsz, s=model.stride.max())  # check img_size

        # Multi-GPU disabled, incompatible with .half() https://github.com/ultralytics/yolov5/issues/99
        # if device.type != 'cpu' and torch.cuda.device_count() > 1:
        #     model = nn.DataParallel(model)

    # Configure, 保留
    model.eval()
    nc = 20
    iouv = torch.linspace(0.5, 0.95, 10).to(device)  # iou vector for mAP@0.5:0.95
    niou = iouv.numel()

    # Dataloader
    if not training:
        #创建一个大小为 (1, 3, imgsz, imgsz) 的全零张量，表示一个批次的单张图像（3 通道，大小为 imgsz x imgsz），并将其分配到指定设备（如 GPU 或 CPU）
        img = torch.zeros((1, 3, imgsz, imgsz), device=device)  # init img
        #如果设备类型不是 CPU，将图像转换为半精度（fp16）或浮点精度（fp32），然后运行模型一次。这是为了初始化模型的一些内部状态。否则不做任何处理。
        _ = model(img) if device.type != 'cpu' else None  # run once
        #根据任务类型（test 或 val）选择测试或验证数据的路径。
        path = data  # path to val/test images
        #创建数据加载器。create_dataloader 函数使用指定的参数创建一个数据加载器，用于加载测试或验证数据。
        dataloader = create_dataloader(path, imgsz, batch_size, model.stride.max(), opt,
                                       hyp=None, augment=False, cache=False, pad=0.5, rect=True)[0]

    seen = 0
    names = ['aeroplane', 'bicycle', 'bird', 'boat', 'bottle', 'bus', 'car', 'cat', 'chair', 'cow',
             'diningtable', 'dog', 'horse', 'motorbike', 'person', 'pottedplant', 'sheep', 'sofa', 'train', 'tvmonitor']
    #设置用于显示的表头格式 s 和一些初始评估指标变量（精度 p、召回率 r、F1 分数 f1、平均精度 mp、mAP@0.5 map50、mAP@0.5:0.95 map 等）
    s = ('%20s' + '%12s' * 6) % ('Class', 'Images', 'Targets', 'P', 'R', 'mAP@.5', 'mAP@.5:.95')
    p, r, f1, mp, mr, map50, map, t0, t1 = 0., 0., 0., 0., 0., 0., 0., 0., 0.
    #初始化一些用于存储评估结果的列表（jdict、stats、ap、ap_class）。
    jdict, stats, ap, ap_class = [], [], [], []

    # quantize_forward & freeze
    if quantize:
        model.quantize(quant_type, num_bits, e_bits)
        for batch_i, (img, targets, paths, shapes) in enumerate(tqdm(dataloader, desc=s)):
            img = img.to(device, non_blocking=True)
            img = img.float()  # uint8 to fp16/32
            img /= 255.0  # 0 - 255 to 0.0 - 1.0
            targets = targets.to(device)
            nb, _, height, width = img.shape  # batch size, channels, height, width
            whwh = torch.Tensor([width, height, width, height]).to(device)

            # Disable gradients
            with torch.no_grad():
                # Run model
                model.quantize_forward(img)  # inference and training outputs
            
        model.freeze()


    # 图像处理，应该可以全部保留
    for batch_i, (img, targets, paths, shapes) in enumerate(tqdm(dataloader, desc=s)):
        img = img.to(device, non_blocking=True)
        img = img.float()  # uint8 to fp16/32
        img /= 255.0  # 0 - 255 to 0.0 - 1.0
        targets = targets.to(device)
        nb, _, height, width = img.shape  # batch size, channels, height, width
        whwh = torch.Tensor([width, height, width, height]).to(device)

        # Disable gradients
        with torch.no_grad():
            # Run model
            t = time_synchronized()
            if not quantize:
                inf_out, train_out = model(img)  # inference and training outputs
            else:
                inf_out, train_out = model.quantize_inference(img)
            t0 += time_synchronized() - t

            # Run NMS
            t = time_synchronized()
            output = non_max_suppression(inf_out, conf_thres=conf_thres, iou_thres=iou_thres, merge=merge)
            t1 += time_synchronized() - t

        # Statistics per image
        for si, pred in enumerate(output):
            labels = targets[targets[:, 0] == si, 1:]
            nl = len(labels)
            tcls = labels[:, 0].tolist() if nl else []  # target class
            seen += 1

            if pred is None:
                if nl:
                    stats.append((torch.zeros(0, niou, dtype=torch.bool), torch.Tensor(), torch.Tensor(), tcls))
                continue

            # Append to text file
            if save_txt:
                gn = torch.tensor(shapes[si][0])[[1, 0, 1, 0]]  # normalization gain whwh
                txt_path = str(out / Path(paths[si]).stem)
                pred[:, :4] = scale_coords(img[si].shape[1:], pred[:, :4], shapes[si][0], shapes[si][1])  # to original
                for *xyxy, conf, cls in pred:
                    xywh = (xyxy2xywh(torch.tensor(xyxy).view(1, 4)) / gn).view(-1).tolist()  # normalized xywh
                    with open(txt_path + '.txt', 'a') as f:
                        f.write(('%g ' * 5 + '\n') % (cls, *xywh))  # label format

            # Clip boxes to image bounds
            clip_coords(pred, (height, width))

            # Assign all predictions as incorrect
            correct = torch.zeros(pred.shape[0], niou, dtype=torch.bool, device=device)
            if nl:
                detected = []  # target indices
                tcls_tensor = labels[:, 0]

                # target boxes
                tbox = xywh2xyxy(labels[:, 1:5]) * whwh

                # Per target class
                for cls in torch.unique(tcls_tensor):
                    ti = (cls == tcls_tensor).nonzero(as_tuple=False).view(-1)  # prediction indices
                    pi = (cls == pred[:, 5]).nonzero(as_tuple=False).view(-1)  # target indices

                    # Search for detections
                    if pi.shape[0]:
                        # Prediction to target ious
                        ious, i = box_iou(pred[pi, :4], tbox[ti]).max(1)  # best ious, indices

                        # Append detections
                        for j in (ious > iouv[0]).nonzero(as_tuple=False):
                            d = ti[i[j]]  # detected target
                            if d not in detected:
                                detected.append(d)
                                correct[pi[j]] = ious[j] > iouv  # iou_thres is 1xn
                                if len(detected) == nl:  # all targets already located in image
                                    break

            # Append statistics (correct, conf, pcls, tcls)
            stats.append((correct.cpu(), pred[:, 4].cpu(), pred[:, 5].cpu(), tcls))

        # Plot images
        if batch_i < 1:
            f = Path(save_dir) / ('test_batch%g_gt.jpg' % batch_i)  # filename
            plot_images(img, targets, paths, str(f), names)  # ground truth
            f = Path(save_dir) / ('test_batch%g_pred.jpg' % batch_i)
            plot_images(img, output_to_target(output, width, height), paths, str(f), names)  # predictions

    # Compute statistics
    stats = [np.concatenate(x, 0) for x in zip(*stats)]  # to numpy
    if len(stats) and stats[0].any():
        p, r, ap, f1, ap_class = ap_per_class(*stats)
        p, r, ap50, ap = p[:, 0], r[:, 0], ap[:, 0], ap.mean(1)  # [P, R, AP@0.5, AP@0.5:0.95]
        mp, mr, map50, map = p.mean(), r.mean(), ap50.mean(), ap.mean()
        nt = np.bincount(stats[3].astype(np.int64), minlength=nc)  # number of targets per class
    else:
        nt = torch.zeros(1)

    # Print results
    pf = '%20s' + '%12.3g' * 6  # print format
    print(pf % ('all', seen, nt.sum(), mp, mr, map50, map))

    # Print results per class
    if verbose and nc > 1 and len(stats):
        for i, c in enumerate(ap_class):
            print(pf % (names[c], seen, nt[c], p[i], r[i], ap50[i], ap[i]))

    # Print speeds
    t = tuple(x / seen * 1E3 for x in (t0, t1, t0 + t1)) + (imgsz, imgsz, batch_size)  # tuple
    if not training:
        print('Speed: %.1f/%.1f/%.1f ms inference/NMS/total per %gx%g image at batch-size %g' % t)

    return map50
    # Return results
    '''
    maps = np.zeros(nc) + map
    for i, c in enumerate(ap_class):
        maps[c] = ap[i]
    return (mp, mr, map50, map, *(loss.cpu() / len(dataloader)).tolist()), maps, t
    '''


if __name__ == '__main__':
    VOC_ROOT = "/workspace/I/wangyuanzhe0/Cam_Model_TF/yolo_v5/data/datasets/voc0712test"
    WEIGHT_ROOT = "/workspace/I/wangyuanzhe0/Cam_Model_TF/yolo_v5/weights/best.pt"
    parser = argparse.ArgumentParser(prog='ptq_one.py')
    parser.add_argument('--weights', nargs='+', type=str, default=WEIGHT_ROOT, help='model.pt path(s)')
    parser.add_argument('--data', type=str, default=VOC_ROOT, help='*.data path')
    parser.add_argument('--batch-size', type=int, default=128, help='size of each image batch')
    parser.add_argument('--img-size', type=int, default=640, help='inference size (pixels)')
    parser.add_argument('--conf-thres', type=float, default=0.001, help='object confidence threshold')
    parser.add_argument('--iou-thres', type=float, default=0.65, help='IOU threshold for NMS')
    parser.add_argument('--save-json', action='store_true', help='save a cocoapi-compatible JSON results file')
    parser.add_argument('--task', default='val', help="'val', 'test', 'study'")
    parser.add_argument('--device', default='cuda', help='cuda device, i.e. 0 or 0,1,2,3 or cpu')
    parser.add_argument('--single-cls', action='store_true', help='treat as single-class dataset')
    parser.add_argument('--augment', action='store_true', help='augmented inference')
    parser.add_argument('--merge', action='store_true', help='use Merge NMS')
    parser.add_argument('--verbose', action='store_true', help='report mAP by class')
    parser.add_argument('--save-txt', action='store_true', help='save results to *.txt')
    parser.add_argument('--quantize-inference', action='store_true', help='quantize-inference for yolo')
    opt = parser.parse_args()
    opt.save_json = False
    #opt.data = check_file(opt.data)  # check file
    print(opt)

    #初始化全局字典
    gol._init()

    #输出文件初始化
    ptq_result_path = 'ptq_result/'

    excel_path = ptq_result_path+'/'+'Yolo_v5'+'.xlsx'
    txt_path = ptq_result_path+'/'+'Yolo_v5'+'.txt'

    workbook = openpyxl.Workbook()
    ft = open(txt_path,'w')

    full_state_dict = torch.load(WEIGHT_ROOT)
    full_model = Yolo_v5()
    quantized_state_dict = full_model.state_dict()
    modify_quant_state(quantized_state_dict,full_state_dict)
    full_model.load_state_dict(quantized_state_dict)
    full_model.eval()
    print('Finished loading model!')

    if opt.device == 'cuda':
        full_model.cuda()
    
    full_mAP = test(opt.data,opt.weights,opt.batch_size,opt.img_size,opt.conf_thres,
                    opt.iou_thres,opt.save_json,opt.single_cls,opt.augment,opt.verbose,
                    full_model,quantize=False,quant_type='INT',num_bits=8,e_bits=3)
    
    Mac, Param, layer, par_ratio, macs_ratio = extract_ratio('Yolo_v5','voc0712test')
    full_names = []
    full_params = []

    for name, param in full_model.named_parameters():
        if 'conv' in name or 'fc' in name or 'cv2.weight' in name or 'cv3.weight' in name:
            full_names.append(name)
            full_params.append(param.data.cpu())

    print('Full_mAP:',full_mAP)

    quant_type_list = ['INT','POT','FLOAT'] #,'POT','FLOAT'
    title_list = []
    js_macs_list = []
    js_param_list = []
    ptq_mAP_list = []
    mAP_loss_list = []

    for quant_type in quant_type_list:
        # FIXME:
        num_bit_list = numbit_list(quant_type)
        # num_bit_list = [8,16]
        # 对一个量化类别，只需设置一次bias量化表
        # int由于位宽大，使用量化表开销过大，直接_round即可
        
        for num_bits in num_bit_list:
            # FIXME:
            e_bit_list = ebit_list(quant_type,num_bits)
            # e_bit_list = [5] 
            for e_bits in e_bit_list:
                if quant_type != 'INT':
                    bias_list = build_bias_list(quant_type,num_bits,e_bits)
                    gol.set_value(bias_list, is_bias=True)

                if quant_type == 'FLOAT':
                    title = '%s_%d_E%d' % (quant_type, num_bits, e_bits)
                else:
                    title = '%s_%d' % (quant_type, num_bits)
                print('\n'+'Yolo_v5'+': PTQ: '+title)
                title_list.append(title)

                # 设置量化表
                if quant_type != 'INT':
                    plist = build_list(quant_type, num_bits, e_bits)
                    gol.set_value(plist)
                # quantize_inference
                print('Evaluation of quantize_inference model:')

                model_ptq = Yolo_v5()
                model_ptq.load_state_dict(quantized_state_dict)
                model_ptq.eval()
                if opt.device == 'cuda':
                    model_ptq.cuda()
                
                ptq_mAP = test(opt.data,opt.weights,opt.batch_size,opt.img_size,opt.conf_thres,
                               opt.iou_thres,opt.save_json,opt.single_cls,opt.augment,opt.verbose,
                               model_ptq,quantize=True,quant_type=quant_type,num_bits=num_bits,e_bits=e_bits)
                ptq_mAP_list.append(ptq_mAP)
                mAP_loss = (full_mAP - ptq_mAP)/full_mAP
                mAP_loss_list.append(mAP_loss)

                #将量化后分布反量化到全精度相同的scale
                model_ptq.fakefreeze()
                # 获取计算量/参数量下的js-div
                js_macs = 0.
                js_param = 0.
                for name, param in model_ptq.named_parameters():
                    if 'conv' not in name and 'fc' not in name and 'cv2.weight' not in name and 'cv3.weight' not in name:
                        continue
                        
                    prefix = name.rsplit('.',1)[0]
                    layer_idx = layer.index(prefix)
                    name_idx = full_names.index(name)

                    layer_idx = layer.index(prefix)
                    ptq_param = param.data.cpu()

                    # FIXME:
                    js = js_div(ptq_param,full_params[name_idx])
                    js = js.item()
                    if js < 0.:
                        js = 0.

                    #TODO: 对于敏感度加权，可以用prefix作为key索引
                    js_macs = js_macs + js * macs_ratio[layer_idx]
                    js_param = js_param + js * par_ratio[layer_idx]


                js_macs_list.append(js_macs)
                js_param_list.append(js_param)

                print(f"ptq mAP: {ptq_mAP:.10f}")

                # 使用旧式的字符串格式化，并设置显示小数点后十位
                print(f"{title}: js_macs: {js_macs:.10f}  js_param: {js_param:.10f} mAP_loss: {mAP_loss:.10f}")
                
    # 写入xlsx
    print("Check here!")
    worksheet = workbook.active
    worksheet.cell(row=1,column=1,value='FP32-mAP')
    worksheet.cell(row=1,column=2,value=full_mAP)
    worksheet.cell(row=1,column=3,value='Mac')
    worksheet.cell(row=1,column=4,value=Mac)
    worksheet.cell(row=1,column=5,value='Param')
    worksheet.cell(row=1,column=6,value=Param)

    worksheet.cell(row=3,column=1,value='title')
    worksheet.cell(row=3,column=2,value='js_macs')
    worksheet.cell(row=3,column=3,value='js_param')
    worksheet.cell(row=3,column=4,value='ptq_mAP')
    worksheet.cell(row=3,column=5,value='mAP_loss')
    for i in range(len(title_list)):
        worksheet.cell(row=i+4, column=1, value=title_list[i])
        worksheet.cell(row=i+4, column=2, value=js_macs_list[i])
        worksheet.cell(row=i+4, column=3, value=js_param_list[i])
        worksheet.cell(row=i+4, column=4, value=ptq_mAP_list[i])
        worksheet.cell(row=i+4, column=5, value=mAP_loss_list[i])

    workbook.save(excel_path)
    
    print('Yolo_v5',file=ft)
    print('Full_mAP: %f'%full_mAP,file=ft)
    print('title_list:',file=ft)
    print(title_list,file=ft)
    print('js_macs_list:',file=ft)
    print(js_macs_list, file=ft)
    print('js_param_list:',file=ft)
    print(js_param_list, file=ft)
    print('ptq_mAP_list:',file=ft)
    print(ptq_mAP_list, file=ft)
    print('mAP_loss_list:',file=ft)
    print(mAP_loss_list, file=ft)
    print("\n",file=ft)

    ft.close()