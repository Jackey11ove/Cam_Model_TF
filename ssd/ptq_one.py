"""Adapted from:
    @longcw faster_rcnn_pytorch: https://github.com/longcw/faster_rcnn_pytorch
    @rbgirshick py-faster-rcnn https://github.com/rbgirshick/py-faster-rcnn
    Licensed under The MIT License [see LICENSE for details]
"""

from __future__ import print_function
import torch
import torch.nn as nn
import torch.backends.cudnn as cudnn
from torch.autograd import Variable
from data import VOC_ROOT, VOCAnnotationTransform, VOCDetection, BaseTransform
from data import VOC_CLASSES as labelmap
import torch.utils.data as data
import openpyxl

from ssd_model import SSD
from extract_ratio import extract_ratio
from util import *
from module import *

import gol
import sys
sys.path.append('/workspace/I/wangyuanzhe0/Cam_Model_TF')
import os
import time
import argparse
import numpy as np
import pickle
#import cv2

if sys.version_info[0] == 2:
    import xml.etree.cElementTree as ET
else:
    import xml.etree.ElementTree as ET


def str2bool(v):
    return v.lower() in ("yes", "true", "t", "1")


parser = argparse.ArgumentParser(
    description='Single Shot MultiBox Detector Evaluation')
parser.add_argument('--trained_model',
                    default='/workspace/I/wangyuanzhe0/Cam_Model_TF/ssd/weights/VOC.pth', type=str,
                    help='Trained state_dict file path to open')
parser.add_argument('--save_folder', default='eval/', type=str,
                    help='File path to save results')
parser.add_argument('--confidence_threshold', default=0.01, type=float,
                    help='Detection confidence threshold')
parser.add_argument('--top_k', default=5, type=int,
                    help='Further restrict the number of predictions to parse')
parser.add_argument('--cuda', default=True, type=str2bool,
                    help='Use cuda to train model')
parser.add_argument('--voc_root', default=VOC_ROOT,
                    help='Location of VOC root directory')
parser.add_argument('--cleanup', default=True, type=str2bool,
                    help='Cleanup and remove results files following eval')

args = parser.parse_args()

if not os.path.exists(args.save_folder):
    os.mkdir(args.save_folder)

if torch.cuda.is_available():
    if args.cuda:
        torch.set_default_tensor_type('torch.cuda.FloatTensor')
    if not args.cuda:
        print("WARNING: It looks like you have a CUDA device, but aren't using \
              CUDA.  Run with --cuda for optimal eval speed.")
        torch.set_default_tensor_type('torch.FloatTensor')
else:
    torch.set_default_tensor_type('torch.FloatTensor')

annopath = os.path.join(args.voc_root, 'VOC2007', 'Annotations', '%s.xml')
imgpath = os.path.join(args.voc_root, 'VOC2007', 'JPEGImages', '%s.jpg')
imgsetpath = os.path.join(args.voc_root, 'VOC2007', 'ImageSets',
                          'Main', 'test.txt')
YEAR = '2007'
devkit_path = args.voc_root + 'VOC' + YEAR
dataset_mean = (104, 117, 123)
set_type = 'test'


class Timer(object):
    """A simple timer."""
    def __init__(self):
        self.total_time = 0.
        self.calls = 0
        self.start_time = 0.
        self.diff = 0.
        self.average_time = 0.

    def tic(self):
        # using time.time instead of time.clock because time time.clock
        # does not normalize for multithreading
        self.start_time = time.time()

    def toc(self, average=True):
        self.diff = time.time() - self.start_time
        self.total_time += self.diff
        self.calls += 1
        self.average_time = self.total_time / self.calls
        if average:
            return self.average_time
        else:
            return self.diff


def parse_rec(filename):
    """ Parse a PASCAL VOC xml file """
    tree = ET.parse(filename)
    objects = []
    for obj in tree.findall('object'):
        obj_struct = {}
        obj_struct['name'] = obj.find('name').text
        obj_struct['pose'] = obj.find('pose').text
        obj_struct['truncated'] = int(obj.find('truncated').text)
        obj_struct['difficult'] = int(obj.find('difficult').text)
        bbox = obj.find('bndbox')
        obj_struct['bbox'] = [int(bbox.find('xmin').text) - 1,
                              int(bbox.find('ymin').text) - 1,
                              int(bbox.find('xmax').text) - 1,
                              int(bbox.find('ymax').text) - 1]
        objects.append(obj_struct)

    return objects


def get_output_dir(name, phase):
    """Return the directory where experimental artifacts are placed.
    If the directory does not exist, it is created.
    A canonical path is built using the name from an imdb and a network
    (if not None).
    """
    filedir = os.path.join(name, phase)
    if not os.path.exists(filedir):
        os.makedirs(filedir)
    return filedir


def get_voc_results_file_template(image_set, cls):
    # VOCdevkit/VOC2007/results/det_test_aeroplane.txt
    filename = 'det_' + image_set + '_%s.txt' % (cls)
    filedir = os.path.join(devkit_path, 'results')
    if not os.path.exists(filedir):
        os.makedirs(filedir)
    path = os.path.join(filedir, filename)
    return path


def write_voc_results_file(all_boxes, dataset):
    for cls_ind, cls in enumerate(labelmap):
        print('Writing {:s} VOC results file'.format(cls))
        filename = get_voc_results_file_template(set_type, cls)
        with open(filename, 'wt') as f:
            for im_ind, index in enumerate(dataset.ids):
                dets = all_boxes[cls_ind+1][im_ind]
                if dets == []:
                    continue
                # the VOCdevkit expects 1-based indices
                for k in range(dets.shape[0]):
                    f.write('{:s} {:.3f} {:.1f} {:.1f} {:.1f} {:.1f}\n'.
                            format(index[1], dets[k, -1],
                                   dets[k, 0] + 1, dets[k, 1] + 1,
                                   dets[k, 2] + 1, dets[k, 3] + 1))


def do_python_eval(output_dir='output', use_07=True):
    cachedir = os.path.join(devkit_path, 'annotations_cache')
    aps = []
    # The PASCAL VOC metric changed in 2010
    use_07_metric = use_07
    print('VOC07 metric? ' + ('Yes' if use_07_metric else 'No'))
    if not os.path.isdir(output_dir):
        os.mkdir(output_dir)
    for i, cls in enumerate(labelmap):
        filename = get_voc_results_file_template(set_type, cls)
        rec, prec, ap = voc_eval(
           filename, annopath, imgsetpath, cls, cachedir,
           ovthresh=0.5, use_07_metric=use_07_metric)
        aps += [ap]
        print('AP for {} = {:.4f}'.format(cls, ap))
        with open(os.path.join(output_dir, cls + '_pr.pkl'), 'wb') as f:
            pickle.dump({'rec': rec, 'prec': prec, 'ap': ap}, f)
    mAP = np.mean(aps)
    print('Mean AP = {:.4f}'.format(np.mean(aps)))
    print('~~~~~~~~')
    print('Results:')
    for ap in aps:
        print('{:.3f}'.format(ap))
    print('{:.3f}'.format(np.mean(aps)))
    print('~~~~~~~~')
    print('')
    print('--------------------------------------------------------------')
    print('Results computed with the **unofficial** Python eval code.')
    print('Results should be very close to the official MATLAB eval code.')
    print('--------------------------------------------------------------')
    return mAP


def voc_ap(rec, prec, use_07_metric=True):
    """ ap = voc_ap(rec, prec, [use_07_metric])
    Compute VOC AP given precision and recall.
    If use_07_metric is true, uses the
    VOC 07 11 point method (default:True).
    """
    if use_07_metric:
        # 11 point metric
        ap = 0.
        for t in np.arange(0., 1.1, 0.1):
            if np.sum(rec >= t) == 0:
                p = 0
            else:
                p = np.max(prec[rec >= t])
            ap = ap + p / 11.
    else:
        # correct AP calculation
        # first append sentinel values at the end
        mrec = np.concatenate(([0.], rec, [1.]))
        mpre = np.concatenate(([0.], prec, [0.]))

        # compute the precision envelope
        for i in range(mpre.size - 1, 0, -1):
            mpre[i - 1] = np.maximum(mpre[i - 1], mpre[i])

        # to calculate area under PR curve, look for points
        # where X axis (recall) changes value
        i = np.where(mrec[1:] != mrec[:-1])[0]

        # and sum (\Delta recall) * prec
        ap = np.sum((mrec[i + 1] - mrec[i]) * mpre[i + 1])
    return ap


def voc_eval(detpath,
             annopath,
             imagesetfile,
             classname,
             cachedir,
             ovthresh=0.5,
             use_07_metric=True):
    """rec, prec, ap = voc_eval(detpath,
                           annopath,
                           imagesetfile,
                           classname,
                           [ovthresh],
                           [use_07_metric])
Top level function that does the PASCAL VOC evaluation.
detpath: Path to detections
   detpath.format(classname) should produce the detection results file.
annopath: Path to annotations
   annopath.format(imagename) should be the xml annotations file.
imagesetfile: Text file containing the list of images, one image per line.
classname: Category name (duh)
cachedir: Directory for caching the annotations
[ovthresh]: Overlap threshold (default = 0.5)
[use_07_metric]: Whether to use VOC07's 11 point AP computation
   (default True)
"""
# assumes detections are in detpath.format(classname)
# assumes annotations are in annopath.format(imagename)
# assumes imagesetfile is a text file with each line an image name
# cachedir caches the annotations in a pickle file
# first load gt
    if not os.path.isdir(cachedir):
        os.mkdir(cachedir)
    cachefile = os.path.join(cachedir, 'annots.pkl')
    # read list of images
    with open(imagesetfile, 'r') as f:
        lines = f.readlines()
    imagenames = [x.strip() for x in lines]
    if not os.path.isfile(cachefile):
        # load annots
        recs = {}
        for i, imagename in enumerate(imagenames):
            recs[imagename] = parse_rec(annopath % (imagename))
            if i % 100 == 0:
                print('Reading annotation for {:d}/{:d}'.format(
                   i + 1, len(imagenames)))
        # save
        print('Saving cached annotations to {:s}'.format(cachefile))
        with open(cachefile, 'wb') as f:
            pickle.dump(recs, f)
    else:
        # load
        with open(cachefile, 'rb') as f:
            recs = pickle.load(f)

    # extract gt objects for this class
    class_recs = {}
    npos = 0
    for imagename in imagenames:
        R = [obj for obj in recs[imagename] if obj['name'] == classname]
        bbox = np.array([x['bbox'] for x in R])
        difficult = np.array([x['difficult'] for x in R]).astype(np.bool_)
        det = [False] * len(R)
        npos = npos + sum(~difficult)
        class_recs[imagename] = {'bbox': bbox,
                                 'difficult': difficult,
                                 'det': det}

    # read dets
    detfile = detpath.format(classname)
    with open(detfile, 'r') as f:
        lines = f.readlines()
    if any(lines) == 1:

        splitlines = [x.strip().split(' ') for x in lines]
        image_ids = [x[0] for x in splitlines]
        confidence = np.array([float(x[1]) for x in splitlines])
        BB = np.array([[float(z) for z in x[2:]] for x in splitlines])

        # sort by confidence
        sorted_ind = np.argsort(-confidence)
        sorted_scores = np.sort(-confidence)
        BB = BB[sorted_ind, :]
        image_ids = [image_ids[x] for x in sorted_ind]

        # go down dets and mark TPs and FPs
        nd = len(image_ids)
        tp = np.zeros(nd)
        fp = np.zeros(nd)
        for d in range(nd):
            R = class_recs[image_ids[d]]
            bb = BB[d, :].astype(float)
            ovmax = -np.inf
            BBGT = R['bbox'].astype(float)
            if BBGT.size > 0:
                # compute overlaps
                # intersection
                ixmin = np.maximum(BBGT[:, 0], bb[0])
                iymin = np.maximum(BBGT[:, 1], bb[1])
                ixmax = np.minimum(BBGT[:, 2], bb[2])
                iymax = np.minimum(BBGT[:, 3], bb[3])
                iw = np.maximum(ixmax - ixmin, 0.)
                ih = np.maximum(iymax - iymin, 0.)
                inters = iw * ih
                uni = ((bb[2] - bb[0]) * (bb[3] - bb[1]) +
                       (BBGT[:, 2] - BBGT[:, 0]) *
                       (BBGT[:, 3] - BBGT[:, 1]) - inters)
                overlaps = inters / uni
                ovmax = np.max(overlaps)
                jmax = np.argmax(overlaps)

            if ovmax > ovthresh:
                if not R['difficult'][jmax]:
                    if not R['det'][jmax]:
                        tp[d] = 1.
                        R['det'][jmax] = 1
                    else:
                        fp[d] = 1.
            else:
                fp[d] = 1.

        # compute precision recall
        fp = np.cumsum(fp)
        tp = np.cumsum(tp)
        rec = tp / float(npos)
        # avoid divide by zero in case the first detection matches a difficult
        # ground truth
        prec = tp / np.maximum(tp + fp, np.finfo(np.float64).eps)
        ap = voc_ap(rec, prec, use_07_metric)
    else:
        rec = -1.
        prec = -1.
        ap = -1.

    return rec, prec, ap


def test_net(save_folder, quantize_flag, net, quant_type, num_bits, e_bits, cuda, dataset, transform, top_k,
             im_size=300, thresh=0.05):
    num_images = len(dataset)
    # all detections are collected into:
    #    all_boxes[cls][image] = N x 5 array of detections in
    #    (x1, y1, x2, y2, score)
    all_boxes = [[[] for _ in range(num_images)]
                 for _ in range(len(labelmap)+1)]

    # timers
    _t = {'im_detect': Timer(), 'misc': Timer()}
    output_dir = get_output_dir('VOC_output', set_type)
    det_file = os.path.join(output_dir, 'detections.pkl')


    if quantize_flag == True:
        net.quantize(quant_type, num_bits, e_bits)
        net.eval()
        for i in range(num_images):
            im, gt, h, w = dataset.pull_item(i)
            x = Variable(im.unsqueeze(0))
            if args.cuda:
                x = x.cuda()
            _t['im_detect'].tic()

            net.quantize_forward(x)
        net.freeze()

    for i in range(num_images):
        im, gt, h, w = dataset.pull_item(i)

        x = Variable(im.unsqueeze(0))
        if args.cuda:
            x = x.cuda()
        _t['im_detect'].tic()
        
        # Most important part
        if quantize_flag == False:
            detections = net.forward(x).data
        else:
            detections = net.quantize_inference(x).data


        detect_time = _t['im_detect'].toc(average=False)

        # skip j = 0, because it's the background class
        for j in range(1, detections.size(1)):
            dets = detections[0, j, :]
            mask = dets[:, 0].gt(0.).expand(5, dets.size(0)).t()
            dets = torch.masked_select(dets, mask).view(-1, 5)
            if dets.size(0) == 0:
                continue
            boxes = dets[:, 1:]
            boxes[:, 0] *= w
            boxes[:, 2] *= w
            boxes[:, 1] *= h
            boxes[:, 3] *= h
            scores = dets[:, 0].cpu().numpy()
            cls_dets = np.hstack((boxes.cpu().numpy(),
                                  scores[:, np.newaxis])).astype(np.float32,
                                                                 copy=False)
            all_boxes[j][i] = cls_dets

        #print('im_detect: {:d}/{:d} {:.3f}s'.format(i + 1,num_images, detect_time))

    with open(det_file, 'wb') as f:
        pickle.dump(all_boxes, f, pickle.HIGHEST_PROTOCOL)

    print('Evaluating detections')
    mAP = evaluate_detections(all_boxes, output_dir, dataset)
    return mAP


def evaluate_detections(box_list, output_dir, dataset):
    write_voc_results_file(box_list, dataset)
    mAP = do_python_eval(output_dir)
    return mAP


def modify_quant_state(quantized_state_dict, full_state_dict):
    name_mapping = {
        'vgg.0.weight':'vgg_upto_conv4_3.conv1.weight',
        'vgg.2.weight':'vgg_upto_conv4_3.conv2.weight',
        'vgg.5.weight':'vgg_upto_conv4_3.conv3.weight',
        'vgg.7.weight':'vgg_upto_conv4_3.conv4.weight',
        'vgg.10.weight':'vgg_upto_conv4_3.conv5.weight',
        'vgg.12.weight':'vgg_upto_conv4_3.conv6.weight',
        'vgg.14.weight':'vgg_upto_conv4_3.conv7.weight',
        'vgg.17.weight':'vgg_upto_conv4_3.conv8.weight',
        'vgg.19.weight':'vgg_upto_conv4_3.conv9.weight',
        'vgg.21.weight':'vgg_upto_conv4_3.conv10.weight',
        'vgg.24.weight':'vgg_upto_fc7.conv11.weight',
        'vgg.26.weight':'vgg_upto_fc7.conv12.weight',
        'vgg.28.weight':'vgg_upto_fc7.conv13.weight',
        'vgg.31.weight':'vgg_upto_fc7.conv14.weight',
        'vgg.33.weight':'vgg_upto_fc7.conv15.weight',
        
        'vgg.0.bias':'vgg_upto_conv4_3.conv1.bias',
        'vgg.2.bias':'vgg_upto_conv4_3.conv2.bias',
        'vgg.5.bias':'vgg_upto_conv4_3.conv3.bias',
        'vgg.7.bias':'vgg_upto_conv4_3.conv4.bias',
        'vgg.10.bias':'vgg_upto_conv4_3.conv5.bias',
        'vgg.12.bias':'vgg_upto_conv4_3.conv6.bias',
        'vgg.14.bias':'vgg_upto_conv4_3.conv7.bias',
        'vgg.17.bias':'vgg_upto_conv4_3.conv8.bias',
        'vgg.19.bias':'vgg_upto_conv4_3.conv9.bias',
        'vgg.21.bias':'vgg_upto_conv4_3.conv10.bias',
        'vgg.24.bias':'vgg_upto_fc7.conv11.bias',
        'vgg.26.bias':'vgg_upto_fc7.conv12.bias',
        'vgg.28.bias':'vgg_upto_fc7.conv13.bias',
        'vgg.31.bias':'vgg_upto_fc7.conv14.bias',
        'vgg.33.bias':'vgg_upto_fc7.conv15.bias',

        'L2Norm.weight':'L2Norm.weight',

        'extras.0.weight':'extra_conv6_2.conv16.weight',
        'extras.1.weight':'extra_conv6_2.conv17.weight',
        'extras.2.weight':'extra_conv7_2.conv18.weight',
        'extras.3.weight':'extra_conv7_2.conv19.weight',
        'extras.4.weight':'extra_conv8_2.conv20.weight',
        'extras.5.weight':'extra_conv8_2.conv21.weight',
        'extras.6.weight':'extra_conv9_2.conv22.weight',
        'extras.7.weight':'extra_conv9_2.conv23.weight',

        'extras.0.bias':'extra_conv6_2.conv16.bias',
        'extras.1.bias':'extra_conv6_2.conv17.bias',
        'extras.2.bias':'extra_conv7_2.conv18.bias',
        'extras.3.bias':'extra_conv7_2.conv19.bias',
        'extras.4.bias':'extra_conv8_2.conv20.bias',
        'extras.5.bias':'extra_conv8_2.conv21.bias',
        'extras.6.bias':'extra_conv9_2.conv22.bias',
        'extras.7.bias':'extra_conv9_2.conv23.bias',

        'loc.0.weight':'loc_layers.0.conv_loc.weight',
        'loc.1.weight':'loc_layers.1.conv_loc.weight',
        'loc.2.weight':'loc_layers.2.conv_loc.weight',
        'loc.3.weight':'loc_layers.3.conv_loc.weight',
        'loc.4.weight':'loc_layers.4.conv_loc.weight',
        'loc.5.weight':'loc_layers.5.conv_loc.weight',

        'loc.0.bias':'loc_layers.0.conv_loc.bias',
        'loc.1.bias':'loc_layers.1.conv_loc.bias',
        'loc.2.bias':'loc_layers.2.conv_loc.bias',
        'loc.3.bias':'loc_layers.3.conv_loc.bias',
        'loc.4.bias':'loc_layers.4.conv_loc.bias',
        'loc.5.bias':'loc_layers.5.conv_loc.bias',

        'conf.0.weight':'conf_layers.0.conv_conf.weight',
        'conf.1.weight':'conf_layers.1.conv_conf.weight',
        'conf.2.weight':'conf_layers.2.conv_conf.weight',
        'conf.3.weight':'conf_layers.3.conv_conf.weight',
        'conf.4.weight':'conf_layers.4.conv_conf.weight',
        'conf.5.weight':'conf_layers.5.conv_conf.weight',

        'conf.0.bias':'conf_layers.0.conv_conf.bias',
        'conf.1.bias':'conf_layers.1.conv_conf.bias',
        'conf.2.bias':'conf_layers.2.conv_conf.bias',
        'conf.3.bias':'conf_layers.3.conv_conf.bias',
        'conf.4.bias':'conf_layers.4.conv_conf.bias',
        'conf.5.bias':'conf_layers.5.conv_conf.bias',
    }

    for old_name, new_name in name_mapping.items():
        if old_name in full_state_dict:
            quantized_state_dict[new_name] = full_state_dict[old_name]
        else:
            print("Warning: model not found in the full precision model state_dict")


if __name__ == '__main__':
    # load net
    num_classes = len(labelmap) + 1                      # +1 for background

    #初始化全局字典
    gol._init()

    #输出文件初始化
    ptq_result_path = 'ptq_result/'

    excel_path = ptq_result_path+'/'+'SSD'+'.xlsx'
    txt_path = ptq_result_path+'/'+'SSD'+'.txt'

    workbook = openpyxl.Workbook()
    ft = open(txt_path,'w')

    #全精度模型
    full_model = SSD(num_classes)            # initialize SSD
    full_state_dict = torch.load(args.trained_model)
    quantized_state_dict = full_model.state_dict()
    modify_quant_state(quantized_state_dict,full_state_dict)
    full_model.load_state_dict(quantized_state_dict)
    full_model.eval()
    print('Finished loading full model!')

    # load data
    dataset = VOCDetection(args.voc_root, [('2007', set_type)],
                           BaseTransform(300, dataset_mean),
                           VOCAnnotationTransform())
    if args.cuda:
        full_model = full_model.cuda()
        cudnn.benchmark = True 

    # full_inference:
    print('Evaluation of full_inference model:')
    quantize_flag = False
    full_mAP = test_net(args.save_folder, quantize_flag, full_model,'INT', 8, 3, args.cuda, dataset,
             BaseTransform(300, dataset_mean), args.top_k, 300,
             thresh=args.confidence_threshold)
    
    Mac, Param, layer, par_ratio, macs_ratio = extract_ratio('SSD','voc2007')
    full_names = []
    full_params = []

    for name, param in full_model.named_parameters():
        if 'conv' in name or 'fc' in name:
            full_names.append(name)
            full_params.append(param.data.cpu())    
    

    quant_type_list = ['INT','POT','FLOAT'] #,'POT','FLOAT'
    title_list = []
    js_macs_list = []
    js_param_list = []
    ptq_mAP_list = []
    mAP_loss_list = []

    for quant_type in quant_type_list:
        # FIXME:
        num_bit_list = numbit_list(quant_type)
        # num_bit_list = [8,16]
        # 对一个量化类别，只需设置一次bias量化表
        # int由于位宽大，使用量化表开销过大，直接_round即可
        
        for num_bits in num_bit_list:
            # FIXME:
            e_bit_list = ebit_list(quant_type,num_bits)
            # e_bit_list = [5] 
            for e_bits in e_bit_list:
                if quant_type != 'INT':
                    bias_list = build_bias_list(quant_type,num_bits,e_bits)
                    gol.set_value(bias_list, is_bias=True)

                if quant_type == 'FLOAT':
                    title = '%s_%d_E%d' % (quant_type, num_bits, e_bits)
                else:
                    title = '%s_%d' % (quant_type, num_bits)
                print('\n'+'SSD'+': PTQ: '+title)
                title_list.append(title)

                # 设置量化表
                if quant_type != 'INT':
                    plist = build_list(quant_type, num_bits, e_bits)
                    gol.set_value(plist)
                # quantize_inference
                print('Evaluation of quantize_inference model:')

                model_ptq = SSD(num_classes)
                model_ptq.load_state_dict(quantized_state_dict)
                model_ptq.eval()
                if args.cuda:
                    model_ptq = model_ptq.cuda()
                    cudnn.benchmark = True
                quantize_flag = True
                ptq_mAP = test_net(args.save_folder, quantize_flag, model_ptq, quant_type, num_bits, e_bits, args.cuda, 
                                   dataset, BaseTransform(300, dataset_mean), args.top_k, 300,
                                   thresh=args.confidence_threshold)
                ptq_mAP_list.append(ptq_mAP)
                mAP_loss = (full_mAP - ptq_mAP)/full_mAP
                mAP_loss_list.append(mAP_loss)

                #将量化后分布反量化到全精度相同的scale
                model_ptq.fakefreeze()
                # 获取计算量/参数量下的js-div
                js_macs = 0.
                js_param = 0.
                for name, param in model_ptq.named_parameters():
                    if 'conv' not in name and 'fc' not in name:
                        continue
                        
                    prefix = name.rsplit('.',1)[0]
                    layer_idx = layer.index(prefix)
                    name_idx = full_names.index(name)

                    layer_idx = layer.index(prefix)
                    ptq_param = param.data.cpu()

                    # FIXME:
                    js = js_div(ptq_param,full_params[name_idx])
                    js = js.item()
                    if js < 0.:
                        js = 0.

                    #TODO: 对于敏感度加权，可以用prefix作为key索引
                    js_macs = js_macs + js * macs_ratio[layer_idx]
                    js_param = js_param + js * par_ratio[layer_idx]


                js_macs_list.append(js_macs)
                js_param_list.append(js_param)

                print(f"ptq mAP: {ptq_mAP:.10f}")

                # 使用旧式的字符串格式化，并设置显示小数点后十位
                print(f"{title}: js_macs: {js_macs:.10f}  js_param: {js_param:.10f} mAP_loss: {mAP_loss:.10f}")                
                
    # 写入xlsx
    print("Check here!")
    worksheet = workbook.active
    worksheet.cell(row=1,column=1,value='FP32-mAP')
    worksheet.cell(row=1,column=2,value=full_mAP)
    worksheet.cell(row=1,column=3,value='Mac')
    worksheet.cell(row=1,column=4,value=Mac)
    worksheet.cell(row=1,column=5,value='Param')
    worksheet.cell(row=1,column=6,value=Param)

    worksheet.cell(row=3,column=1,value='title')
    worksheet.cell(row=3,column=2,value='js_macs')
    worksheet.cell(row=3,column=3,value='js_param')
    worksheet.cell(row=3,column=4,value='ptq_mAP')
    worksheet.cell(row=3,column=5,value='mAP_loss')
    for i in range(len(title_list)):
        worksheet.cell(row=i+4, column=1, value=title_list[i])
        worksheet.cell(row=i+4, column=2, value=js_macs_list[i])
        worksheet.cell(row=i+4, column=3, value=js_param_list[i])
        worksheet.cell(row=i+4, column=4, value=ptq_mAP_list[i])
        worksheet.cell(row=i+4, column=5, value=mAP_loss_list[i])

    workbook.save(excel_path)
    
    print('SSD',file=ft)
    print('Full_mAP: %f'%full_mAP,file=ft)
    print('title_list:',file=ft)
    print(title_list,file=ft)
    print('js_macs_list:',file=ft)
    print(js_macs_list, file=ft)
    print('js_param_list:',file=ft)
    print(js_param_list, file=ft)
    print('ptq_mAP_list:',file=ft)
    print(ptq_mAP_list, file=ft)
    print('mAP_loss_list:',file=ft)
    print(mAP_loss_list, file=ft)
    print("\n",file=ft)

    ft.close()